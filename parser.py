import re
import time
from openpyxl import load_workbook
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options


def get_driver():
    opt = Options()
    opt.add_argument("--disable-infobars")
    opt.add_argument("start-maximized")
    opt.add_argument("--disable-extensions")
    # Pass the argument 1 to allow and 2 to block
    opt.add_experimental_option("prefs", { \
        "profile.default_content_setting_values.media_stream_mic": 1,
        "profile.default_content_setting_values.media_stream_camera": 1,
        "profile.default_content_setting_values.geolocation": 1,
        "profile.default_content_setting_values.notifications": 1
    })

    driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=opt)  # pip install webdriver_manager
    return driver


def poisk_reorg_i_glav_vrach_po_ogrn(ogrn="1025006174852", vivod_iteracii=1):
    time.sleep(1)
    elem = driver.find_element_by_xpath("/html/body/div/section/div/div/form/label/input")
    elem.clear()
    elem.send_keys(ogrn)
    time.sleep(1.2)
    t1 = driver.find_element_by_xpath("/html/body/div/section/div/div/div/div/a/div[1]")
    t2 = driver.find_element_by_xpath("/html/body/div/section/div/div/div/div/a/div[2]")
    t3 = driver.find_element_by_xpath("/html/body/div/section/div/div/div/div/a/div[3]")
    temp = ["ul", "glav_vrach", "status"]
    temp[0] = t1.text
    if re.findall(r'\d+', str(t3.text)):
        temp[1] = t2.text
        temp[2] = "ДЕЙСТВУЮЩАЯ"
    else:
        temp[1] = t3.text
        temp[2] = t2.text

    print(t1.text, "Добавлена", vivod_iteracii)
    return temp


def get_ogrn_by_xl_file(dest_filename='словарь с объединениями2.xlsx', sheet="Лист2"):
    ogrn_in_excel = []
    wb = load_workbook(filename=dest_filename)
    sheet_ranges = wb[sheet]
    column_k = sheet_ranges['K']
    for cell in column_k:
        if cell.value is not None:
            ogrn_in_excel.append(cell.value)
    return ogrn_in_excel


def save(dest_filename='словарь с объединениями2.xlsx', sheet="Лист2"):
    wb = load_workbook(filename=dest_filename)
    sheet_ranges = wb[sheet]
    for iterator in range(-1, len(new_info)):
        if iterator == -1:
            pass
        else:

            cell = sheet_ranges.cell(row=iterator + 2, column=15)
            cell.value = new_info[iterator][0]
            cell = sheet_ranges.cell(row=iterator + 2, column=16)
            cell.value = new_info[iterator][1]
            cell = sheet_ranges.cell(row=iterator + 2, column=17)
            cell.value = new_info[iterator][2]

    wb.save(dest_filename)
    print("Записано в", dest_filename)


def obrabotat_vse_ogrn():
    for i in range(1, len(ogrn_in_excel)):
        try:
            new_info.append(poisk_reorg_i_glav_vrach_po_ogrn(ogrn_in_excel[i], i))
        except:
            input("\n Введите капчу, затем нажмите любоую кнопку ")
            new_info.append(poisk_reorg_i_glav_vrach_po_ogrn(ogrn_in_excel[i], i))


driver = get_driver()
driver.get("https://www.rusprofile.ru/")
assert "Проверка и анализ" in driver.title
ogrn_in_excel = get_ogrn_by_xl_file()
print("\033[33m {}".format("Получили ОГРН из файла"))
new_info = []
obrabotat_vse_ogrn()

save()
driver.close()